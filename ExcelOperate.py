import pandas as pd
import xlrd
import openpyxl
import tempfile

from linkfix import instagram_link_fix


class ExcelOperate:
    """
    功能:
        1. 加载xlsx文件或者xls文件，将其转换为data数据或者pandas的数据
        2. 单文件去重
        3. 两文件相对去重
    """
    # 从Excel工作簿转换来的数据
    data: dict
    # 从data转换到pandas的数据
    df: pd.core.frame.DataFrame

    def __init__(self, excel_name=None, excel_file=None):
        if excel_file:
            self.workbook, self.EXCEL_TYPE = self.load_excel(excel_name, excel_file)

            if self.EXCEL_TYPE == "XLRD":
                ExcelOperate.data = self.read_with_xlrd(wb=self.workbook)

            if self.EXCEL_TYPE == "OPENPYXL":
                ExcelOperate.data = self.read_with_openpyxl(wb=self.workbook)

            ExcelOperate.df = self.to_pandas(data=ExcelOperate.data)

    def load_excel(self, excel_name, excel_file):
        """
        加载不同的Excel文件类型
        返回工作簿对象
        
        returns:
            wb (openpyxl or xlrd): Excel工作簿对象
            XLRD or OPENPYXL (str): 工作簿类型
        """
        if (excel_name.split(".")[1] == "xls"):
            # 因为openpyxl库只支持从本地路径读取Excel文件
            # 故创建本地临时文件供openpyxl读取
            with tempfile.NamedTemporaryFile(delete=False) as f:
                f.write(excel_file.read())
                excel_file = f.name

            wb = xlrd.open_workbook(excel_file)
    
            XLRD = "XLRD"
            print('Using "xlrd."')
    
            return (wb, XLRD)
        
        if (excel_name.split(".")[1] == "xlsx"):
            wb = openpyxl.load_workbook(excel_file)
    
            OPENPYXL = "OPENPYXL"
            print('Using "openpyxl."')
    
            return (wb, OPENPYXL)
    
    def read_with_xlrd(self, wb: xlrd.book.Book) -> dict:
        """
        读取xls格式的Excel工作簿的数据
    
        返回的数据格式:
        {
            "Sheet1": [
                {
                    "Name": [
                        "10.14 added\\ue05e\u71e5",
                        "metejnica"
                    ]
                },
                {
                    "Channel Link": [
                        "empty:",
                        "https://www.instagram.com/metejnica/"
                    ]
                }
            ],
            "Sheet2": [
                {
                    "Name": [
                        "3.15 added \\ue05e\u71e5",
                        "lafamilialatorre"
                    ]
                },
                {
                    "Channel Link": [
                        "empty:",
                        "https://www.instagram.com/lafamilialatorre/"
                    ]
                }
            ],
            ...
        }
        """
        # 储存工作簿数据
        data = {}
    
        for sheet_name in wb.sheet_names():
            table = wb.sheet_by_name(sheet_name)
            nrows, ncols = table.nrows, table.ncols
    
            # table为空则跳过
            if (nrows + ncols == 0):
                continue
    
            print(f'读取 "{sheet_name}", 有 {nrows} 行 {ncols} 列.')
    
            # 添加表格
            data.update({sheet_name: []})
    
            # 向表格添加表头
            column_names = []
            for col in range(ncols):
                column_name = str(table.row(0)[col]).replace("text:", "").replace("'", "")
                column_names.append(column_name)
                data[sheet_name].append({column_name: []})
    
            # 在表头下追加数据
            for row in range(1, nrows):
                for col in range(ncols):
                    value = str(table.row(row)[col]).replace("text:", "").replace("'", "")
                    data[sheet_name][col][column_names[col]].append(value)
    
        return data
    
    def read_with_openpyxl(self, wb: openpyxl.workbook.workbook.Workbook) -> dict:
        """
        读取xlsx格式的Excel工作簿的数据
    
        返回的数据格式:
            与函数 read_with_xlrd 返回的格式相同
        """
        # 储存工作簿数据
        data = {}
    
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            nrows, ncols = sheet.max_row, sheet.max_column
    
            # 如果Sheet为0行0列则跳过
            if (nrows + ncols == 0):
                continue
    
            print(f'读取 "{sheet_name}", 有 {nrows} 行 {ncols} 列.')
    
            # 添加表格
            data.update({sheet_name: []})
    
            # 向表格添加表头
            column_names = []
            for col in range(1, ncols+1):
                column_name = sheet.cell(row=1, column=col).value
                column_names.append(column_name)
                data[sheet_name].append({column_name: []})
    
            # 在表格下追加数据
            for row in range(1, nrows+1):
                for col in range(1, ncols+1):
                    value = str(sheet.cell(row=row, column=col).value)
                    data[sheet_name][col-1][column_names[col-1]].append(value)
        
        return data
    
    def to_pandas(self, data: dict) -> pd.core.frame.DataFrame:
        """
        将Excel工作簿中的数据转换为pandas数据格式
        注意:
            1. 工作簿中所有sheet都将合并
            2. 合并后可能会出现重复数据(在原Excel工作簿中已经重复的数据)
    
        pandas数据格式:
        {
            "key": ["value1", "value2", ...],
            ...
        }
        """
        pandas_data = {}
    
        # 将read_with_xlrd中得到的数据转换为pandas数据
        for data_key, data_value in data.items():
            for sheet in data_value:
                for sheet_key, sheet_value in sheet.items():
                    if (sheet_key not in pandas_data.keys()):
                        pandas_data.update({sheet_key: []})
                    for values in sheet.values():
                        for value in values:
                            pandas_data[sheet_key].append(value)
    
        df = pd.DataFrame(pandas_data)
    
        return df
    
    def deduplicates(self, df_compare: pd.core.frame.DataFrame, df_dest: pd.core.frame.DataFrame) -> pd.core.frame.DataFrame:
        """
        功能:
            先由deduplicate函数对两个文件进行处理
            去除df_dest中与df_compare重复的行(以链接为唯一判断标准)
    
        returns:
            pandas格式数据
        """
        df_compare = self.deduplicate(df=df_compare)
        df_dest = self.deduplicate(df=df_dest)
    
        links = []
    
        # TODO
        """
            1. 保存df_compare中的所有链接
            2. 添加data，遍历df_dest中所有的链接，if value not in df_compare，则添加
        """
        # 保存df_compare中所有的链接
        for key, values in df_compare.items():
            links = [value for value in values if "https" in value]
    
        # 储存df_dest数据
        data = {}
        # 重复标志
        DEDUPLICATE_FLAG = False
    
        # 添加表头
        for key, values in df_dest.items():
            data.update({key: []})
    
        # 忽略与df_compare重复的内容(以链接作为唯一标准)
        for row in range(len(df_dest)):
            DEDUPLICATE_FLAG = False
    
            for value in df_dest.values[row]:
                if ("https" in value and value in links):
                    DEDUPLICATE_FLAG = True
    
            if not DEDUPLICATE_FLAG:
                for key, values in df_dest.items():
                    data[key].append(values[row])
    
        df = pd.DataFrame(data)
    
        return df
    
    def deduplicate(self, df: pd.core.frame.DataFrame) -> pd.core.frame.DataFrame:
        """
        功能:
            1. 修复链接
            2. 单独文件去重
    
        returns:
            pandas格式的数据
        """
        # 修复链接后的pandas数据
        data_link_fixed = {}
    
        # 修复链接
        for key in df.keys():
            data_link_fixed.update({key: []})
            for value in df[key]:
                # 修正链接
                if "instagram" in value:
                    value = instagram_link_fix(link=value)
                # 跳过表头
                if key != value:
                    data_link_fixed[key].append(value)
    
        # 去重后的pandas数据
        data_deduplicated = {}
        # 数据中不包含链接的行不需要去重
        DEDUPLICATE_FLAG = False

        # 空表格处理
        if not [len(values) for values in data_link_fixed.values()]:
            return pd.DataFrame(
                {
                    "None": ["None"]
                }
            )

        # 获取最大行数与最大列数
        ncols = len(data_link_fixed.keys())
        nrows = max([len(values) for values in data_link_fixed.values()])
    
        # 添加表头
        for key in data_link_fixed.keys():
            data_deduplicated.update({key: []})
    
        # 去重操作
        for row in range(nrows):
            # 重置flag为False
            DEDUPLICATE_FLAG = False
    
            # 检查数据是否出现过，以链接为唯一判断标准
            for key in data_deduplicated.keys():
                value = data_link_fixed[key][row]
    
                if ("https" in value and value in data_deduplicated[key]):
                    DEDUPLICATE_FLAG = True
    
            # 将不重复的数据追加到data_deduplicated中
            if not DEDUPLICATE_FLAG:
                for key in data_deduplicated.keys():
                    value = data_link_fixed[key][row]
                    data_deduplicated[key].append(value)
    
        df = pd.DataFrame(data_deduplicated)
    
        return df


if __name__ == "__main__":
    pass
